"""
CSV to Pretty Excel Converter for LeetCode Problems

This script:
1. Reads the LeetCode problems CSV file
2. Cleans and formats the data
3. Creates a styled Excel spreadsheet with:
   - Color-coded difficulty levels
   - Proper number formatting
   - Clickable hyperlinks
   - Clean column widths
   - Freeze panes for headers
   - Conditional formatting
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.hyperlink import Hyperlink
import argparse

# Color definitions
COLOR_EASY = 'C6EFCE'  # Light green
COLOR_MEDIUM = 'FFEB9C'  # Light yellow
COLOR_HARD = 'FFC7CE'  # Light red
HEADER_COLOR = '4472C4'  # Blue
HEADER_FONT_COLOR = 'FFFFFF'  # White
LINK_COLOR = '0563C1'  # Blue for hyperlinks

def clean_csv_data(df):
    """Clean and normalize the CSV data"""
    # Remove percentage signs from Acceptance column
    if 'Acceptance' in df.columns:
        df['Acceptance'] = df['Acceptance'].astype(str).str.replace('%', '').astype(float)
    
    # Clean other potential numeric columns
    numeric_cols = ['Frequency', 'ID']
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
    
    return df

def create_styled_excel(df, output_file):
    """Create a styled Excel file from the DataFrame"""
    # Create a workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "LeetCode Problems"
    
    # Write the DataFrame to the worksheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # Create styles
    header_font = Font(bold=True, color=HEADER_FONT_COLOR)
    header_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
    center_aligned = Alignment(horizontal='center')
    link_font = Font(color=LINK_COLOR, underline='single')
    
    # Apply header styles
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_aligned
        cell.border = thin_border
    
    # Apply styles to data cells and make links clickable
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            if cell.column_letter in ['A', 'C', 'D', 'E']:  # ID, Acceptance, Difficulty, Frequency
                cell.alignment = center_aligned
            
            # Make LeetCode links clickable
            if cell.column_letter == 'F' and cell.value:  # Leetcode Question Link column
                cell.hyperlink = cell.value
                cell.font = link_font
                cell.style = "Hyperlink"
    
    # Apply conditional formatting for difficulty
    if 'Difficulty' in df.columns:
        difficulty_col = None
        for idx, col in enumerate(ws.iter_cols(min_row=1, max_row=1), 1):
            if col[0].value == 'Difficulty':
                difficulty_col = idx
                break
        
        if difficulty_col:
            # Easy - green
            ws.conditional_formatting.add(
                f'D2:D{ws.max_row}',
                CellIsRule(operator='equal', formula=['"Easy"'], 
                          fill=PatternFill(start_color=COLOR_EASY, end_color=COLOR_EASY, fill_type="solid"))
            )
            # Medium - yellow
            ws.conditional_formatting.add(
                f'D2:D{ws.max_row}',
                CellIsRule(operator='equal', formula=['"Medium"'], 
                          fill=PatternFill(start_color=COLOR_MEDIUM, end_color=COLOR_MEDIUM, fill_type="solid"))
            )
            # Hard - red
            ws.conditional_formatting.add(
                f'D2:D{ws.max_row}',
                CellIsRule(operator='equal', formula=['"Hard"'], 
                          fill=PatternFill(start_color=COLOR_HARD, end_color=COLOR_HARD, fill_type="solid"))
            )
    
    # Format Acceptance as percentage
    if 'Acceptance' in df.columns:
        for cell in ws['C'][1:]:  # Skip header
            cell.number_format = '0.0"%'
    
    # Format Frequency to 3 decimal places
    if 'Frequency' in df.columns:
        for cell in ws['E'][1:]:  # Skip header
            cell.number_format = '0.000'
    
    # Set column widths
    column_widths = {
        'A': 8,    # ID
        'B': 50,   # Title
        'C': 12,   # Acceptance
        'D': 12,   # Difficulty
        'E': 12,   # Frequency
        'F': 60    # Leetcode Question Link
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Save the workbook
    wb.save(output_file)
    print(f"Successfully created styled Excel file with clickable links: {output_file}")

def main():
    parser = argparse.ArgumentParser(description='Convert LeetCode CSV to styled Excel with clickable links')
    parser.add_argument('input_csv', help='Path to input CSV file')
    parser.add_argument('output_xlsx', help='Path for output Excel file')
    
    args = parser.parse_args()
    
    try:
        # Read and clean the data
        df = pd.read_csv(args.input_csv)
        df = clean_csv_data(df)
        
        # Create styled Excel file with clickable links
        create_styled_excel(df, args.output_xlsx)
        
    except Exception as e:
        print(f"Error processing file: {e}")
        print("Please check your input file and try again.")

if __name__ == '__main__':
    main()